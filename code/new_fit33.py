import numpy as np
import matplotlib.pyplot as plt
import pickle

import PDL1_Integral_Inputs_ATFL_influx_slope_us_problem as problem
from scipy.optimize import least_squares, differential_evolution,shgo,dual_annealing,minimize,Bounds,BFGS

import datetime
import shutil
import inspect
cur_timestamp = datetime.datetime.strftime(datetime.datetime.now(),'%Y%m%d_%H%M%S')
export_dir = 'fig_exp.d/new_fit/'+'exp_'+cur_timestamp+'_'
log_dir = 'log.d/new_fit/'+'log_'+cur_timestamp+'_'
def dubf():
    pass 
shutil.copyfile(inspect.getsourcefile(dubf),log_dir+'new_fit.py.txt')
shutil.copyfile(inspect.getsourcefile(problem.problem_input),log_dir+'problem.py.txt')

N_total_pat = 9

def get_exp_data():
    import openpyxl
    exp_data = np.full([N_total_pat, problem.N_FitTreatment, problem.N_ResultField],np.nan)
    exp_error = np.full([N_total_pat, problem.N_FitTreatment, problem.N_ResultField],np.nan)

    # Cell Phenotype Percentage Data
    wb1 = openpyxl.load_workbook("../Fig5H+Fig5S3_corrected.xlsx")

    ws_data = list(wb1)[0]
    head_rows_data = {problem.ResultField.apop:2, problem.ResultField.CD154:10, problem.ResultField.CD163:19}
    rows_data = {problem.FitTreatment.Control:0, problem.FitTreatment.aCSF1R:1, problem.FitTreatment.aPD1:2, problem.FitTreatment.dual:3}
    cols_data = [chr(x+67) for x in range(9)]

    ws_err = list(wb1)[1]
    head_rows_err = {problem.ResultField.apop:2, problem.ResultField.CD154:9, problem.ResultField.CD163:16}
    rows_err = {problem.FitTreatment.Control:0, problem.FitTreatment.aCSF1R:1, problem.FitTreatment.aPD1:2, problem.FitTreatment.dual:3}
    cols_err = [chr(x+67) for x in range(9)]

    for fld in head_rows_data.keys():
        for treat in rows_data.keys():
            for pat in range(N_total_pat):
                exp_data[pat, treat, fld] = ws_data[cols_data[pat]+str(head_rows_data[fld]+rows_data[treat])].value
                exp_error[pat, treat, fld] = ws_err[cols_err[pat]+str(head_rows_err[fld]+rows_err[treat])].value
    
    # Cytokine Concentration Data
    wb2 =  openpyxl.load_workbook("../Fig5F+Fig5S2+Fig3G.xlsx")
    ws = list(wb2)[0]
    head_rows_data = {problem.FitTreatment.Control:3, problem.FitTreatment.aCSF1R:11, problem.FitTreatment.aPD1:24, problem.FitTreatment.dual:37}
    rows_data = {problem.ResultField.TNFa:0, problem.ResultField.IFNg:1, problem.ResultField.TGFb:2, problem.ResultField.IL10:3}
    cols_data = [chr(x+67) for x in range(9)]
    # FIXME There is no error data for control in this xlsx.
    head_rows_err = {problem.FitTreatment.Control:17, problem.FitTreatment.aCSF1R:17, problem.FitTreatment.aPD1:30, problem.FitTreatment.dual:43}
    rows_err = {problem.ResultField.TNFa:0, problem.ResultField.IFNg:1, problem.ResultField.TGFb:2, problem.ResultField.IL10:3}
    cols_data = [chr(x+67) for x in range(9)]

    for fld in rows_data.keys():
        for treat in head_rows_data.keys():
            for pat in range(N_total_pat):
                exp_data[pat, treat, fld] = ws[cols_data[pat]+str(head_rows_data[treat]+rows_data[fld])].value
                exp_error[pat, treat, fld] = ws[cols_err[pat]+str(head_rows_err[treat]+rows_err[fld])].value
    assert np.all(exp_error>0)
    return exp_data, exp_error

# def get_residual_matrix_mask():
#     rei

param_fixed={
    'xinit':[1,2],
    'detailed':False,
    # 'k5':1,
    # 'k8':1,
    # 'wTGFb':1,
    # 'wIL10':1,
    # 'wTNFa':1,
    # 'wIFNg':1,
    'rTGFb':1,
    'rIL10':1,
    'rTNFa':1,
    'rIFNg':1,
    # 'e':3,
    # 'r1': np.log(2)/(23/24),
    # 'r1': np.log(2)/(3),
    # 'k12':1/(72/24),
    # 'k21_2_k12':2,
    # 'l2':0,
    'nua_ratio':0,
    'sPD_o_ref':1,
    'cCSF_o_ref':1,
    'n0':1,
    'Tinh_CSF': 1/(1+1e6*0.1/398.48/1),
    'Tinh_PD':1/(1+1e6*1/143597/1.45),

    'KTGFb':1,
    'KIL10':1,
    'KTNFa':1,
    'KIFNg':1,
    'tau':1,
    # 'nM':3,
    'nM':1e6,
    'KA':None,
    'KU':None,
    'Tt':30,
    'rCD154':None,
    'ATFLin':None,
    'nM':None,
    'DPDL1os_variation':None,

    # 'k3':None,
    # 'k4':None,
    # 'k5':None,
    # 'k6':None,
    # 'k7':None,
    # 'k8':None,
    # 'k9':None,
    # 'k10':None,
    # 'k11':None,
    # 'wTGFb':None,
    # 'wIL10':None,
    # 'wTNFa':None,
    # 'wIFNg':None,
    # 'wm2':None,
    # 'wm1':None,
    # 'Km2':None,
    # 'Km1':None,

    'l1':6e-3,
    # 'l2':4e-3,
    # 'l2':0.4,
    'l3':4e-3,
    'k5':0,
    'k8':0,
    # 'kact_max':3.2
    'kact_max':1
}

## Home grown initials
param_shared={
    # 'r1': (np.log(2)/(23/24), [0,1]),
    # 'e': (30,[0,1000]),
    # 'e':(3,[0,30]),
    'e':(3,[0.3,30]),
    
    # 'r2':(np.log(2)/2,[0.1,3]),
    # 'r2':(np.log(2)/2,[np.log(2)/3,np.log(2)/2]),
    'r2_2_r1':(1,[1,1.1]),
    # 'r2':(1,[0.1,10]),
    # 'k12':(1/(72/24),[0.01,0.1]),
    'k12':(0.05,[0.01,0.1]),
    # 'k21_2_k12':(2,[1,10]),
    'k21_2_k12':(100,[1,300]),
    # 'l2': (0.1,[0,6]),
    'l2': (0.2,[0.2,0.7]),
    # 'l2': (0,[0,1]),
    # 'l3': (0,[0,2]),
    # 'l2': (0.1,[0,0.1]),
    # 'l3': (0.1,[0,0.1]),
    # 'k3':(100,[0,500]),
    # 'k4':(1,[0,500]),
    # 'k5':(1,[0,2000]),
    # 'k6':(100,[0,500]),
    # 'k3':(100,[1e-2,500]),
    # 'k6':(100,[1e-2,500]),
    'k3':(10,[1e-2,100]),
    'k6':(10,[1e-2,100]),
    # 'k7':(1,[0,500]),
    # 'k8':(1,[0,2000]),
    # 'k9':(1,[0,100]),
    # 'k10':(1,[0,100]),
    # 'k11':(1,[0,300]),
    # 'k9':(1,[0,100]),
    # 'k10':(1,[0,100]),
    # 'k11':(1,[0,300]),
    'k9':(1,[1e-2,100]),
    'k10':(1,[1e-2,10]),
    'k11':(1,[1e-2,10]),
    # 'kact_max':(0.05,[0.05,100]),
    # 'kact_max':(1,[0.01,4]),
    # 'kact_max':(1,[0.01,5]),
    # 'DPDL_o_sPDL':(0.1,[0,10]),
    # 'sPD_o_ref':(0.5,[0,1]),
    # 'cCSF_o_ref':(0.5,[0,1]),
    # 'Km2':(1,[1e-4,10]),
    # 'Km1':(1,[1e-4,10]),
    'Km2':(1,[1e-2,10]),
    'Km1':(1,[1e-2,10]),
    # 'Dag':(1,[0,30]),
    # 'KTGFb':(1000,[0,50000]),
    # 'KIL10':(1000,[0,50000]),
    # 'KTNFa':(1,[0,5000]),
    # 'KIFNg':(1,[0,5000]),

    # 'wTGFb':(0.1,[0,1]),
    # 'wIL10':(0.1,[0,1]),
    # 'wTNFa':(0.1,[0,1]),
    # 'wIFNg':(0.1,[0,1]),

    'wTGFb':(1,[0,1]),
    'wIL10':(1,[0,1]),
    'wTNFa':(1,[0,1]),
    'wIFNg':(1,[0,1]),
    # 'wm2':(1,[1e-4,10]),
    # 'wm1':(1,[1e-4,10]),
    # 'wm2':(1e-4,[1e-4,1]),
    'wm2':(1,[0,1]),
    # 'wm1':(1e-4,[1e-4,1]),
    'wm1':(1,[0,1]),
    # 'nua_ratio':(0,[0,0.1]),
    # 'na_n_nn_0':(0.5,[0,0.5])
    # 'na_n_nn_0':(1,[1,10]),
    'na_n_nn_0':(1,[1,8]),
    # 'rCD154':(1,[1/4,1]),
    # 'KA':(1,[0,5]),
    # 'nM':(100,[3,100]),
    # 'KU':(1,[0,4]),
    'Dag':(1,[1e-2,30]),
    'r1': (np.log(2)/2, [np.log(2)/3,np.log(2)/2]),
    'DPDL_o_sPDL_base':(1,[1e-2,10]),
    'k4_base':(1e-2,[1e-2,10]),
    'k7_base':(1e-2,[1e-2,10]),

    
    

}

param_subtype={
    # 'r1': (np.log(2)/2, [np.log(2)/3,np.log(2)/2]),
    # 'l1': (0,[0,0.1]),
    # 'l1': (0.1,[0,0.1]),
    # 'na_n_nn_0':(2,[0.5,6]),
    # 'na_n_nn_0':(1,[0.5,6]),
    # 'na_n_nn_0':(1,[1,10]),
    # 'DPDL_o_sPDL':(1,[1e-2,10]),
    # 'DPDL_o_sPDL':(1,[1e-4,10]),
    # 'DPDL_o_sPDL':(1,[1e-2,10]),
    'DPDL_o_sPDL_factor':(1,[1,2]),
    # 'Dag':(1,[0,30]),
    # 'Dag':(1,[1e-2,30]),
    # 'k4':(0,[0,500]),
    # 'k5':(0,[0,2000]),
    # 'k7':(0,[0,500]),
    # 'k8':(0,[0,2000]),
    'k4_factor':(1,[1,2]),
    # 'k5':(1e-2,[1e-2,1]),
    'k7_factor':(1,[1,2]),
    # 'k8':(1e-2,[1e-2,1]),
}

param_individual={
    'cCSF_o_DCSF':(10,[0,20]),
    # 'DPDL_o_sPDL':(1,[1e-2,10]),
    # 'na_n_nn_0':(3,[0.5,6]),
    # 'nua_ratio':(0,[0,0.1]),
    # 'rTGFb':(1,[0,100]),
    # 'rIL10':(1,[0,100]),
    # 'rTNFa':(1,[0,100]),
    # 'rIFNg':(1,[0,100]),

    
}

assert(all([x not in list(param_fixed) for x in list(param_shared)]))
assert(all([x not in list(param_fixed) for x in list(param_individual)]))
assert(all([x not in list(param_fixed) for x in list(param_subtype)]))
assert(all([x not in list(param_individual) for x in list(param_shared)]))
assert(all([x not in list(param_individual) for x in list(param_subtype)]))
assert(all([x not in list(param_shared) for x in list(param_subtype)]))


exp_data, exp_error = get_exp_data()
original_exp_error = exp_error.copy()
## Averaging error bar for all treatment and patients with each single data field.
# for fld in problem.ResultField:
for fld in [problem.ResultField.apop, problem.ResultField.CD154]:
    exp_error[:,:,fld] = np.mean(exp_error[:,:,fld])

# patient2fit = [0,1,2,3,4,5,6,7,8]
# patient2fit = [1]
# patient2fit = [2]
# patient2fit=[0,2]
# patient2fit = [3,5]
# patient2fit = [6,8]
# patient2fit = [0,2,3,5,6,8]
patient2fit = [0,2,4,5,6,8]
# patient2fit = [1,7]
# patient2fit = [1,4,7]
# patient2check = [2,5,8]
patient2check = []
# patient2check = [1,4,7]
# patient2check = [3]

total_subtype_list = [[0,1,2],[3,4,5],[6,7,8]]
N_total_subtype = len(total_subtype_list)
subtype_of_total_pat = []
for kp in range(N_total_pat):
    for ks in range(N_total_subtype):
        if kp in total_subtype_list[ks]:
            subtype_of_total_pat.append(ks)
            break
print(subtype_of_total_pat)


pat_exp_data, pat_exp_error = exp_data[patient2fit,:,:], exp_error[patient2fit,:,:]
N_patfit = len(patient2fit)
key_shared = list(param_shared)
key_individual = list(param_individual)
key_subtype = list(param_subtype)
N_shared = len(key_shared)
N_indiv = len(key_individual)
N_P_subtype = len(key_subtype)

subtype_of_pat_fit = np.array(subtype_of_total_pat)[patient2fit]
subtype2fit = np.unique(subtype_of_pat_fit)
N_subtype_fit = len(subtype2fit)
subtype_fit_list = [ [] for x in range(N_subtype_fit)]
for kp,ks in enumerate(subtype_of_pat_fit):
    subtype_fit_list[ks].append(kp)


## Sifting Data
AllPSR_mask = np.full([N_total_pat,problem.N_FitTreatment,problem.N_ResultField],True)
AllPSR_mask[:,problem.FitTreatment.aPD1,problem.ResultField.CD163]=False # Mask out CD163 under aPD1 from sq_residual
# AllPSR_mask[:,problem.FitTreatment.dual, :] = False # Mask out Dual for prediction.
AllPSR_mask[:,:,[problem.ResultField.IL10, problem.ResultField.IFNg, problem.ResultField.TGFb, problem.ResultField.TNFa]]=False
## Mask out contradictory data under aCSF1R
AllPSR_mask[:,problem.FitTreatment.aCSF1R,problem.ResultField.CD154] = False
AllPSR_mask[:,problem.FitTreatment.aCSF1R,problem.ResultField.apop] = False


# AllPSR_mask[[0,3,4,5,6,7],:,problem.ResultField.apop]=False
# AllPSR_mask[[0,5,6,7],:,problem.ResultField.apop]=False
# AllPSR_mask[[0,5,7],:,problem.ResultField.apop]=False
# AllPSR_mask[[0,5,8],:,problem.ResultField.apop]=False
# AllPSR_mask[[0,5],:,problem.ResultField.apop]=False
# AllPSR_mask[[0,5,6],:,problem.ResultField.apop]=False
# AllPSR_mask[[0,5,6,8],:,problem.ResultField.apop]=False
AllPSR_mask[[0,4,5,6,8],:,problem.ResultField.apop]=False
# AllPSR_mask[[0,4,5,8],:,problem.ResultField.apop]=False
# AllPSR_mask[[4,5,7,8],:,problem.ResultField.CD154]=False
AllPSR_mask[[7,8],:,problem.ResultField.CD154]=False
# AllPSR_mask[[5,7,8],:,problem.ResultField.CD154]=False
AllPSR_mask[[0,1,2,3],:,problem.ResultField.CD163]=False

# # TODO temporary for pat #3
AllPSR_mask[2,problem.FitTreatment.Control,problem.ResultField.CD163]=True
AllPSR_mask[2,problem.FitTreatment.dual,problem.ResultField.CD163]=True
# AllPSR_mask[2,:,problem.ResultField.CD154]=False

# # TODO temporary for pat #1
AllPSR_mask[0,problem.FitTreatment.Control,problem.ResultField.CD163]=True
AllPSR_mask[0,problem.FitTreatment.dual,problem.ResultField.CD163]=True

# # TODO temporary for pat #6
AllPSR_mask[5,problem.FitTreatment.aPD1,problem.ResultField.apop]=True
# AllPSR_mask[5,problem.FitTreatment.dual,problem.ResultField.apop]=True

# # TODO temporary for pat #5
AllPSR_mask[4,problem.FitTreatment.aPD1,problem.ResultField.apop]=True
AllPSR_mask[4,problem.FitTreatment.Control,problem.ResultField.apop]=True

# TODO temporary for pat #9
AllPSR_mask[8,problem.FitTreatment.aPD1,problem.ResultField.apop]=True
AllPSR_mask[8,problem.FitTreatment.Control,problem.ResultField.apop]=True
AllPSR_mask[8,problem.FitTreatment.dual,problem.ResultField.apop]=True

# TODO temporary for pat #7
AllPSR_mask[6,problem.FitTreatment.dual,problem.ResultField.apop]=True
AllPSR_mask[6,problem.FitTreatment.Control,problem.ResultField.apop]=True

for kp in patient2fit:
    if exp_data[kp,problem.FitTreatment.Control,problem.ResultField.CD163]>exp_data[kp,problem.FitTreatment.aCSF1R,problem.ResultField.CD163]:
        AllPSR_mask[kp,problem.FitTreatment.Control,problem.ResultField.CD163]=True
        AllPSR_mask[kp,problem.FitTreatment.aCSF1R,problem.ResultField.CD163]=True
    if exp_data[kp,problem.FitTreatment.Control,problem.ResultField.CD163]>exp_data[kp,problem.FitTreatment.dual,problem.ResultField.CD163]:
        AllPSR_mask[kp,problem.FitTreatment.Control,problem.ResultField.CD163]=True
        AllPSR_mask[kp,problem.FitTreatment.dual,problem.ResultField.CD163]=True

PSR_mask = AllPSR_mask[patient2fit,:,:]

# For slope conditions
large_filling = 10000
N_PSR_mask_true_entries = np.count_nonzero(PSR_mask)
filled_unmasked_SR = np.full(N_PSR_mask_true_entries,large_filling)
print(filled_unmasked_SR)


def get_residual(pat_outputs, pat_exp_data, pat_exp_error):
    res =  ((pat_outputs-pat_exp_data)/pat_exp_error)
    # modification of weighting to emphasize apop
    # res[:,:,problem.ResultField.apop]*=2
    return res

def get_sq_residual(pat_outputs, pat_exp_data, pat_exp_error):
    return (get_residual(pat_outputs, pat_exp_data, pat_exp_error))**2
    

def array2param(x):
    param_list = []
    for kp in range(N_patfit):
        pat_param = param_fixed.copy()
        for k1, key in enumerate(key_shared):
            pat_param[key]=x[k1]
        for k3, key in enumerate(key_subtype):
            pat_param[key] = x[N_shared+subtype_of_pat_fit[kp]*N_P_subtype+k3]
        for k2, key in enumerate(key_individual):
            pat_param[key]=x[N_shared+N_P_subtype*N_subtype_fit+N_indiv*kp+k2]
        param_list.append(pat_param)
    return param_list

def param2array(param_list):
    x = np.full(N_shared+N_P_subtype*N_subtype_fit+N_indiv*N_patfit, np.nan)
    for k1,key in enumerate(key_shared):
        x[k1]=param_list[0][key]
    for ks in range(N_subtype_fit):
        for k3, key in enumerate(key_subtype):
            x[N_shared+ks*N_P_subtype+k3] = param_list[subtype_fit_list[ks][0]][key]
    for kp in range(N_patfit):
        for k2, key in enumerate(key_individual):
            x[N_shared+N_P_subtype*N_subtype_fit+N_indiv*kp+k2]=param_list[kp][key]
    return x

def g(x):
    param_list = array2param(x)
    pat_outputs = np.full([N_patfit,problem.N_FitTreatment,problem.N_ResultField],np.nan)
    for kp in range(N_patfit):
        q = problem.problem_input(**param_list[kp])
        # if not all([q.monoclinic_apop,q.monoclinic_CD154]): return filled_unmasked_SR
        monojud_apop = -q.monoclinic_apop*filled_unmasked_SR
        monojud_CD154 = -q.monoclinic_CD154*filled_unmasked_SR

        o = q.block_result
        pat_outputs[kp,:,:]=o.transpose()
    
    unmasked_SR = get_residual(pat_outputs,pat_exp_data,pat_exp_error)
    # print(array2param(x)) 

    with open('dump_x.bin','wb') as f:
        pickle.dump(x,f)
    with open('dump_p.bin','wb') as pf:
        pickle.dump(param_list,pf)

    # return np.ravel(unmasked_SR[:,SR_mask])
    # return np.ravel(unmasked_SR[PSR_mask])
    return np.ravel([unmasked_SR[PSR_mask],monojud_apop,monojud_CD154])

def fit():

    # initialization by default parameters
    param_init={}
    param_upperbound = {}
    param_lowerbound = {}
    for k1, key in enumerate(key_shared):
        param_init[key] = param_shared[key][0]
        param_lowerbound[key] = param_shared[key][1][0]
        param_upperbound[key] = param_shared[key][1][1]
    for k2,key in enumerate(key_individual):
        param_init[key] = param_individual[key][0]
        param_lowerbound[key] = param_individual[key][1][0]
        param_upperbound[key] = param_individual[key][1][1]
    for k3,key in enumerate(key_subtype):
        param_init[key] = param_subtype[key][0]
        param_lowerbound[key] = param_subtype[key][1][0]
        param_upperbound[key] = param_subtype[key][1][1]
    repeat_param_2array = lambda param: param2array([param for x in range(N_patfit)])
    x_init = repeat_param_2array(param_init)
    ### Print any infesible parameters (param_init version)
    print(param_init)
    tmp_p = param_init
    for name in list(param_lowerbound):
        if param_lowerbound[name]>tmp_p[name] or param_upperbound[name]<tmp_p[name]:
            print(name,tmp_p[name])

    init_by_assignment = False
    # init_by_assignment = True
    # initialization by assigned parameters
    if init_by_assignment:
        # By assigning parameters
        # init_p_fname = 'fig_exp.d/new_fit/exp_20210120_003237_res.bin'
        # with open(init_p_fname,'rb') as f:
        #     paramS_init = pickle.load(f)[1]
        ### interim parameters
        init_p_fname = 'dump_p.bin'
        with open(init_p_fname,'rb') as f:
            paramS_init = pickle.load(f)
        ### adjusting parameters if necessary
        for p in paramS_init:
            # p['r1']=np.log(2)/2
            # p['r2']=0.1
            # p['wm1']=0.5
            # p['sPD_o_ref']=0
            # # p['wIFNg']=0
            # # p['wTNFa']=0.1
            # p['wTGFb']=0
            # p['wIL10']=0
            # p['wm2']=0.5
            # # p['sPD_o_ref']=0
            # p['DPDL_o_sPDL']=1
            # p['kact_max']=1
            # p['l2']=0.5
            # p['e']=3
            # p['KA']=1
            # p['nM']=5
            # p['na_n_nn_0']=1
            # p['Dag']=20
            pass
        print(paramS_init)
        ### Load data for multiple patients directly
        x_init = param2array(paramS_init)
        ### Print any infesible parameters
        # tmp_p = paramS_init[0]
        # for name in list(param_lowerbound):
        #     if param_lowerbound[name]>tmp_p[name] or param_upperbound[name]<tmp_p[name]:
        #         print(name,tmp_p[name])
        ### If this is parameter of all pats
        # x_init = param2array(paramS_init)
        ### If this is of one pat
        # x_init = repeat_param_2array(paramS_init[0])
        ## By assigning x
        # with open('dump_x.bin','rb') as f:
        #     x_init = pickle.load(f)
    

    x_lowerbound, x_upperbound = repeat_param_2array(param_lowerbound),repeat_param_2array(param_upperbound)
    x_bound = np.concatenate([x_lowerbound[np.newaxis,:],x_upperbound[np.newaxis,:]],axis=0)


    # x_init[x_init>x_upperbound]=x_upperbound[x_init>x_upperbound]
    # x_init[x_init<x_lowerbound]=x_lowerbound[x_init<x_lowerbound]


    # res = least_squares(g,x_init,bounds = (x_lowerbound,x_upperbound),verbose=2,ftol=1e-5)
    res = least_squares(g,x_init,bounds = (x_lowerbound,x_upperbound),verbose=2)
    # res = least_squares(g,x_init,bounds = (x_lowerbound,x_upperbound),verbose=2,ftol=1e-10,xtol=1e-12)

    ## Using other algorithms
    scipyBounds = Bounds(x_lowerbound,x_upperbound)
    def f(x):
        r=np.sum(g(x)**2)
        print(r)
        return r
    # res = minimize(f,x_init,method='SLSQP',bounds=scipyBounds,options={'disp':True,'maxiter':1000})
    # # # res = dual_annealing(f,bounds = x_bound.T)
    # res = differential_evolution(f,bounds=x_bound.T,tol=0.3)

    res_params = array2param(res.x)
    
    with open(export_dir+'res.bin','wb') as f:
        pickle.dump((res,res_params),f)
    
    return res

# def check(params):

if __name__ == '__main__':
    res = fit()


    



            




        


